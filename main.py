import random
import time

from instabot import Bot
from instaloader import Instaloader, Profile
import shutil
import json
from geopy.geocoders import Nominatim
from multiprocessing import Pool
from openpyxl import load_workbook

# try:
#     shutil.rmtree('config')  # delete folder or instabot will not work
# except FileNotFoundError:
#     print('Config folder doesnt found. Nothing to delete')

my_insta_login = 'elazavetanovikova'
my_insta_pass = 'ItOT8glBEtY'
# my_insta_login = 'anurovan'
# my_insta_pass = '1cdd4e457eb34v5er8'

targer_user = 'jerlis_lives_here'


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< INSTALOADER
def init_instaloader() -> Instaloader:
    loader = Instaloader()
    # loader.interactive_login(my_insta_login)
    loader.login(user=my_insta_login, passwd=my_insta_pass)
    return loader


def protect_from_ban(loader):
    username = ['tomholland2013', 'beyonce']
    profile = Profile.from_username(context=loader.context, username=username[0])
    igtv_posts = profile.get_tagged_posts()


def get_location_by_geotag(users_data):
    useful_data = check_existed_data(users_data)
    target_save_file = 'users_location_data/followees_location_data.json'
    loader = init_instaloader()

    for num_user, user in enumerate(users_data):
        if num_user > 0 and num_user % 10 == 0:
            print('[WARN] Protection from ban')
            time.sleep(random.randint(3, 7))
            protect_from_ban(loader)
        username = user.get('username')
        username_id = user.get('id')
        try:
            profile = Profile.from_username(context=loader.context, username=username)
        except Exception as ex:
            print(f'[ERR] Account ({username}) doesnt exist, {ex}')
            continue
        posts = profile.get_posts()

        current_user = {}
        for num_post, post in enumerate(posts):
            if num_post >= 10:  # search data only in first 10 posts
                break
            try:
                location = post.location
            except Exception as ex:
                print(f'[ERR] Cant get location, {ex}')
                location = None
                continue
            if not location:
                continue
            latitude = str(location.lat)
            longitude = str(location.lng)
            current_user = {
                'username': username,
                'id': username_id,
                'lng': longitude,
                'lat': latitude
            }
            print('[SUCCESS] Location founded')
            break  # stop if find first geotag

        if not current_user:
            current_user = {
                'username': username,
                'id': username_id,
                'lng': 'no_info',
                'lat': 'no_info'
            }
        useful_data.append(current_user)

        with open(target_save_file, 'w', encoding='utf-8') as file:
            file.write(json.dumps(useful_data))
            print(f'[INFO] Writing {num_user + 1} from {len(users_data)}')


def check_existed_data(users_data):
    filename = 'users_location_data/followees_location_data.json'
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            exist_data = json.loads(file.read())
            for e_data in exist_data:
                for u_num, u_data in enumerate(users_data):
                    if e_data.get('id') != u_data.get('id'):
                        continue
                    del users_data[u_num]
            return exist_data
    except Exception as ex:
        return []


def load_json_data():
    with open('followees_data.json', 'r', encoding='utf-8') as file:
        users_data = json.loads(file.read())
    # parts_amount = 1700
    # user_data_parts = [users_data[i * parts_amount:(i + 1) * parts_amount] for i in
    #                    range((len(users_data) + parts_amount - 1) // parts_amount)]
    # p = Pool(processes=len(user_data_parts))
    # p.map(get_location_by_geotag, user_data_parts)
    get_location_by_geotag(users_data)


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< INSTABOT
def init_instabot() -> Bot:
    bot = Bot()
    bot.login(username=my_insta_login, password=my_insta_pass)
    return bot


def write_followers_name_and_info(src_file, target_file):
    bot = init_instabot()
    with open(src_file, 'r', encoding='utf-8') as file:
        followers_dict = json.loads(file.read())

    follower_data = []
    for num, follower in enumerate(followers_dict):
        follower_data.append({
            'username': bot.get_username_from_user_id(follower),
            'id': follower,
            'info': bot.get_user_info(follower)
        })
        print(f'[INFO] Got {num + 1} profile from {len(followers_dict)}')
        if num % 100 == 0:
            print(f'[INFO] Got {num + 1} profile from {len(followers_dict)}. Save data to JSON...')
            with open(target_file, 'w', encoding='utf-8') as file:
                file.write(json.dumps(follower_data))
    with open(target_file, 'w', encoding='utf-8') as file:
        file.write(json.dumps(follower_data))


def write_follows_ids():
    bot = init_instabot()
    followers = bot.get_user_followers(targer_user)
    with open('followers_id.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(followers))

    followers = bot.get_user_following(targer_user)
    with open('followees_id.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(followers))


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< GET LOCATION
def get_location(latitude, longitude):
    if not latitude or not longitude:
        return 'no_info'
    if latitude == 'no_info' or longitude == 'no_info':
        return 'no_info'

    geolocator = Nominatim(user_agent="geoapiExercises")
    try:
        location = geolocator.reverse(f'{latitude},{longitude}', language='en')
    except Exception as ex:
        return 'no_info'
    if not location:
        return 'no_info'
    return location.address


def collect_location():
    useful_data = []
    with open('users_location_data/followees_location_data.json', 'r', encoding='utf-8') as file:
        users_data = json.loads(file.read())
        for num, each in enumerate(users_data):
            try:
                longitude = each.get('lng')
                latitude = each.get('lat')
            except Exception as ex:
                longitude = None
                latitude = None
            useful_data.append({
                'username': each.get('username'),
                'id': each.get('id'),
                'location': get_location(latitude=latitude, longitude=longitude)
            })
            print(f'[INFO] Get {num + 1} user from {len(users_data)}')
            if num % 100 == 0:
                print('Writing to file...')
                with open('followees_location_data.json', 'w', encoding='utf-8') as file:
                    file.write(json.dumps(useful_data))
    with open('followees_location_data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(useful_data))


def collect_all_files():
    with open('new_followees_data.json', 'r', encoding='utf-8') as file:
        followers_list = json.loads(file.read())
    with open('followees_location_data.json', 'r', encoding='utf-8') as file:
        followers_location_list = json.loads(file.read())

    useful_follower_list = []
    for each_follower in followers_list:
        for each_follower_loc in followers_location_list:
            if each_follower_loc.get('id') == each_follower.get('id'):
                location = each_follower_loc.get('location')
                each_follower = {**each_follower, 'location': location}
                useful_follower_list.append(each_follower)
                break
    with open('complete_followees_data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(useful_follower_list))


def write_file_to_excel():
    with open('complete_followees_data.json', 'r', encoding='utf-8') as file:
        python_dict = json.loads(file.read())

        workbook = load_workbook(filename='followees_data.xlsx')
        sheet = 'data'
        excel_column = 1
        excel_row = 1

        # seeking_data
        useful_python_list = []
        for each_dict in python_dict:
            useful_python_list.append({
                'username': each_dict.get('username'),
                'user_id': each_dict.get('id'),
                'location': each_dict.get('location'),
                'full_name': each_dict.get('info').get('full_name'),
                'follower_count': each_dict.get('info').get('follower_count'),
                'following_count': each_dict.get('info').get('following_count'),
                'biography': each_dict.get('info').get('biography'),
                'external_url': each_dict.get('info').get('external_url'),
                'followee_is_follower': each_dict.get('followee_is_follower')
            })

        for key, value in useful_python_list[0].items():
            workbook[sheet].cell(row=excel_row, column=excel_column).value = key
            excel_column += 1

        excel_row = 2
        for data in useful_python_list:
            for num, each in enumerate(data):
                workbook[sheet].cell(row=excel_row, column=num + 1).value = data[each]
            excel_row += 1
        workbook.save('followees_data.xlsx')
        workbook.close()


def check_followees():
    with open('followees_data.json', 'r', encoding='utf-8') as file:
        followees_list = json.loads(file.read())
    with open('followers_data.json', 'r', encoding='utf-8') as file:
        followers_list = json.loads(file.read())

    new_followes_list = []
    for each_followee in followees_list:
        followee_is_follower = 'False'
        for each_follower in followers_list:
            if each_followee.get('id') == each_follower.get('id'):
                followee_is_follower = 'True'
                # new_followee = {**each_followee, 'followee is follower': 'true'}
                # new_followes_list.append(new_followee)
                break
        new_followee = {**each_followee, 'followee_is_follower': followee_is_follower}
        new_followes_list.append(new_followee)

    with open('new_followees_data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(new_followes_list))


if __name__ == '__main__':
    # load_json_data()
    # collect_location()
    # check_followees()
    # collect_all_files()
    write_file_to_excel()
